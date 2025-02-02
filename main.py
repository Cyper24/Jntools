import inquirer
import os
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
from docxtpl import DocxTemplate
from datetime import datetime
import locale

clear = lambda: os.system('cls')
clear()

print("""

░▒▓███████▓▒░ ░▒▓██████▓▒░ ░▒▓██████▓▒░░▒▓███████▓▒░▒▓████████▓▒░▒▓██████▓▒░ ░▒▓██████▓▒░░▒▓█▓▒░       ░▒▓███████▓▒░ 
░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░ ░▒▓█▓▒░  ░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░      ░▒▓█▓▒░        
░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░      ░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░ ░▒▓█▓▒░  ░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░      ░▒▓█▓▒░        
░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒▒▓███▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░ ░▒▓█▓▒░  ░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░       ░▒▓██████▓▒░  
░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░ ░▒▓█▓▒░  ░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░             ░▒▓█▓▒░ 
░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░ ░▒▓█▓▒░  ░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░░▒▓█▓▒░▒▓█▓▒░             ░▒▓█▓▒░ 
░▒▓█▓▒░░▒▓█▓▒░░▒▓██████▓▒░ ░▒▓██████▓▒░░▒▓█▓▒░░▒▓█▓▒░ ░▒▓█▓▒░   ░▒▓██████▓▒░ ░▒▓██████▓▒░░▒▓████████▓▒░▒▓███████▓▒░  
                                                                                 
                                                                                                Made with ❤  By XXiv           
""")

def manifest(kt):
    list = []
    urlkt = "https://jmsgw.jntexpress.id/transportation/trackingDeatil/loading/scan/page"
    urlkt2 = "https://jmsgw.jntexpress.id/transportation/tmsShipment/traceDetail"
    querystring = {"current":"1","size":"20000","shipmentNo":f"{kt}","scanNetworkCode":"SOC999"}
    querystring2 = {"shipmentNo":f"{kt}"}
    print("Please Wait...")
    response = requests.request("GET", urlkt, headers=headers, params=querystring)
    response2 = requests.request("GET", urlkt2, headers=headers, params=querystring2)
    manifest = response.json()
    manifest2 = response2.json()
    fdocx = manifest2["data"]["shipmentDetail"]
    tujuan = fdocx["endName"]
    nopol = fdocx["plateNumber"]
    driver = fdocx["driverName"]
    tgld = fdocx["plannedDepartureTime"]
    locale.setlocale(locale.LC_TIME, 'id_ID')
    xt = datetime.strptime(f'{tgld}', '%Y-%m-%d %H:%M:%S')
    tgal = xt.strftime("%A, %d-%m-%Y / %H.%M")
    doc = DocxTemplate("sj.docx")
    context = {'driver' : driver,
           'nopol': nopol,
            'kt' : kt,
            'tujuan' : tujuan,
            'tgl' : tgal}
    doc.render(context)
    doc.save("sjnew.docx")
    subprocess.Popen(["sjnew.docx"],shell=True)  

    f = manifest["data"]["records"]
    for x in f:
        billCode = x["billCode"]
        packageCode = x["packageCode"]
        final = {'No. Waybill' : billCode,'Kepemilikan No. Bagging' : packageCode}
        list.append(final)
    df = pd.DataFrame(list)
    df.to_excel('temp.xlsx')
    df = pd.read_excel('temp.xlsx')
    df['Kepemilikan No. Bagging'] = df['Kepemilikan No. Bagging'].fillna("-" + kt)
    pivot = df.pivot_table(index=["Kepemilikan No. Bagging"],values=['No. Waybill'],aggfunc=['count'],margins=True, margins_name='Total')
    rows2 = pivot["count"]["No. Waybill"].reset_index()
    wb = Workbook()
    sheet = wb.active
    for r in dataframe_to_rows(rows2):
            sheet.append(r)
        # sheet.delete_cols(idx=1)
    sheet.delete_rows(idx=2)
    sheet.insert_rows(idx=1,amount=2)
    sheet["A1"] = "SOC GATEWAY"
    sheet["A2"] = "OUTGOING SOC GATEWAY TO " + tujuan
    sheet["B3"] = "No Bagging "
    sheet["C3"] = "AWB"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    tc = int(len(sheet['A']))
    bagging = tc - 5
    ecer = sheet["C4"].value - bagging
    sheet["C4"] = ecer
    tr = int(sheet[f"C{tc}"].value)
    total = tr - bagging
    sheet[f"C{tc}"] = total
    bord = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
    for row in range(1,sheet.max_row+1):
                for col in range(1,sheet.max_column+1):
                        cell=sheet.cell(row, col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(color="000000", size=11,bold=True)
                        cell.border = bord
    bold= Font(color="000000", size=14,bold=True)
    sheet["A1"].font = bold
    sheet["A2"].font = bold
    wb.save(filename="manifest.xlsx")
    print("Done")
    subprocess.Popen(["manifest.xlsx"],shell=True)  

def loadunl():
    list2 = []
    url2 = "https://jmsgw.jntexpress.id/transportation/trackingDeatil/loading/scan/list"
    with open("listkt.txt") as f:
        lines = [line.rstrip('\n') for line in f]

    for kt in lines:
        print("=====================================")
        print(kt)
        unload = []
        querystring = {"shipmentNo":f"{kt}"}
        response = requests.request("GET", url2, headers=headers, params=querystring)
        r=response.json()
        for lo in r["data"]:
            if lo["scanNetworkCode"] == "SOC999" and lo["loadingTypeName"] == "1":
                valuelo = lo["scanWaybillNum"]
                valuelo = int(valuelo)
                print(f"Total Load: {valuelo}")
            else:
                    value = 0

        for unl in r["data"]:
            if unl["loadingTypeName"] == "2":
                    valueunl = unl["scanWaybillNum"]
                    valueunl = int(valueunl)
                    unload.append(valueunl)
            else:
                    value = 0
        un = sum(unload)
        print(f"Total Unload: {un}")
        final = {'Kode Tugas' : kt,'Load' : valuelo,'Unload':un}
        list2.append(final)
        df = pd.DataFrame(list2)
    else:
        df.to_csv('jntloadunload.csv')
        subprocess.Popen(["jntloadunload.csv"],shell=True)
        print("Done")

def reportincoming():
    list3 = []
    url3 = "https://jmsgw.jntexpress.id/transportation/tmsBranchTrackingDetail/page"
    tglinput = input("""Input Tanggal
Format Tahun-Bulan-Tanggal // Contoh : 2025-01-26 
: """)
    
    payload = {
        "current": 1,
        "size": 100,
        "startDepartureTime": f"{tglinput} 00:00:00",
        "endDepartureTime": f"{tglinput} 23:59:59",
        "startCode": "SOC999",
        "countryId": "1"
    }
    print("Please Wait...")
    response = requests.request("POST", url3, json=payload, headers=headers)
    r=response.json()
    f = r["data"]["records"]
    for x in f:
        shipmentNo = x["shipmentNo"]
        lineName = x["lineName"]
        loadCount = x["loadCount"]
        actualVehicleTypegroup = x["actualVehicleTypegroup"]
        carrierName = x["carrierName"]
        plannedDepartureTime = x["plannedDepartureTime"]
        actualDepartureTime = x["actualDepartureTime"]
        plannedArrivalTime = x["plannedArrivalTime"]
        actualArrivalTime = x["actualArrivalTime"]
        shifts = x["shifts"]
        final = {'kode Tugas' : shipmentNo,'Rute' : lineName,'Load' : loadCount,'Tipe Armada' : actualVehicleTypegroup,'Vendor' : carrierName,
                 'Perencanaan Waktu Keberangkatan (CUT OFF)' : plannedDepartureTime,'Keberangkatan Aktual Mobil' : actualDepartureTime,'Rencanakan Waktu Kedatangan' : plannedArrivalTime,
                 'Waktu aktual kedatangan Mobil' : actualArrivalTime,'Rit' : shifts}
        list3.append(final)

    df = pd.DataFrame(list3)
    df.to_csv('jntkota.csv')
    print("Done")
    subprocess.Popen(["jntkota.csv"],shell=True)

p_tools = [
    inquirer.List(
        "alat",
        message="Pilih Tools : ",
        choices=["Manifest","Cari Load Unload","Report Incoming","Surat Karantina","Keluar"],
    ),
]

p_tujuan = [
    inquirer.List(
        "tujuan",
        message="Pilih Tujuan",
        choices=["JKT","SEG","SUB","JAT"],
    ),
]

while True:
    url = "https://jmsgw.jntexpress.id/message/messageInfo/popPage"
    payload = {
        "status": 1,
        "size": 10,
        "current": 1,
        "type": 2,
        "batchNos": [],
        "settingUse": 2,
        "countryId": "1"
    }
    choice = input("Enter Auth Token: ")
    headers = {
                "cookie": "HWWAFSESID=a00e27f02785ef49ce5; HWWAFSESTIME=1738201375713",
                "authtoken": f"{choice}",
                "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36 Edg/132.0.0.0",
                "Content-Type": "application/json"
    }
    response = requests.request("POST", url, json=payload, headers=headers)
    if response.status_code == 200:
        for x in range(1000):
            answers = inquirer.prompt(p_tools)
            if answers["alat"] == "Manifest":
                kotug = input("Input Kode Tugas: ")
                manifest(kotug)
            if answers["alat"] == "Cari Load Unload":
                print("Edit listkt.txt Terlebih Dahulu")
                loadunl()
            if answers["alat"] == "Surat Karantina":
                subprocess.call(["C:\\Users\\Mboh\\Documents\\xxiv\\sempak\\dist\\sempak2.exe"])
            if answers["alat"] == "Report Incoming":
                reportincoming()
            if answers["alat"] == "Keluar":
                 quit()

    else:
        print("Enter Valid Authtoken!!")
